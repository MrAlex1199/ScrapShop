import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import os
import platform
import json

# Create folders if they don't exist
for folder in ['data', 'receipts_in', 'receipts_out']:
    if not os.path.exists(folder):
        os.makedirs(folder)


class ScrapShopApp:
    def __init__(self, root):
        self.root = root
        self.root.title("โปรแกรมคำนวณรับซื้อและจำหน่ายของเก่า V6.1")
        self.root.state("zoomed")

        # --- CustomTkinter Style ---
        ctk.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
        # Themes: "blue", "green", "dark-blue"
        ctk.set_default_color_theme("green")

        # --- Initialize prices FIRST before other operations ---
        self.load_prices()
        print(
            f"✅ โหลดราคาสำเร็จ: รับซื้อ {len(self.BUY_PRICES)} รายการ, จำหน่าย {len(self.SELL_PRICES)} รายการ")

        # --- Initialize Thai font ---
        self.thai_font_name = self.register_thai_font()

        # Configure a style for the ttk Treeview to match CTk's theme
        style = ttk.Style(self.root)
        style.theme_use("default")  # Use default theme as a base
        style.configure("Treeview.Heading", font=(
            "TH Sarabun New", 18, "bold"))
        style.configure("Treeview", font=("TH Sarabun New", 16))

        # --- Main Frame ---
        main_frame = ctk.CTkFrame(self.root, corner_radius=10)
        main_frame.pack(fill=ctk.BOTH, expand=True, padx=20, pady=20)

        # --- Notebook (Tabview) ---
        self.notebook = ctk.CTkTabview(main_frame, width=800, height=600)
        self.notebook.pack(fill=ctk.BOTH, expand=True, padx=10, pady=10)

        # Setup Tabs
        self.incoming_tab = self.notebook.add("รับเข้า")
        self.outgoing_tab = self.notebook.add("จำหน่ายออก")
        self.history_tab = self.notebook.add("ประวัติและคงคลัง")

        self._setup_transaction_tab(
            self.incoming_tab, "in", self.BUY_PRICES, self.update_buy_price,
            "ชื่อผู้ขาย:", "ชื่อผู้รับ:"
        )
        self._setup_transaction_tab(
            self.outgoing_tab, "out", self.SELL_PRICES, self.update_sell_price,
            "ชื่อผู้จ่าย:", "ชื่อผู้รับ (เช่น โรงงาน):"
        )
        self.setup_history_tab()

        # Excel and JSON files
        self.incoming_excel = os.path.join(
            'data', 'incoming_scrap_records.xlsx')
        self.outgoing_excel = os.path.join(
            'data', 'outgoing_scrap_records.xlsx')
        self.receipt_history_file = os.path.join(
            'data', 'receipt_history.json')

        if not os.path.exists(self.incoming_excel):
            self.create_excel_file(self.incoming_excel)
        if not os.path.exists(self.outgoing_excel):
            self.create_excel_file(self.outgoing_excel)

        # Load histories
        self.load_excel_history(self.incoming_excel, self.incoming_calc_tree)
        self.load_excel_history(self.outgoing_excel, self.outgoing_calc_tree)
        self.load_receipt_history()

        # Compute inventory
        self.compute_inventory()

        # Current data
        self.current_in_data = None
        self.current_out_data = None

    def load_prices(self):
        """โหลดราคาจากไฟล์ JSON พร้อมการจัดการข้อผิดพลาด"""
        prices_file = 'prices.json'

        # ราคาเริ่มต้นหากไม่มีไฟล์หรือโหลดไม่ได้
        default_buy_prices = {
            "กระดาษลัง": 3.50,
            "กระดาษขาว-ดำ": 6.00,
            "หนังสือพิมพ์": 7.50,
            "ขวดพลาสติกใส (PET)": 13.50,
            "พลาสติกขาวขุ่น/ขวดน้ำ": 11.00,
            "เหล็กหนา": 9.50,
            "เหล็กบาง": 8.00,
            "กระป๋องอลูมิเนียม": 60.00,
            "ทองแดง (เบอร์ 1)": 295.00,
            "สแตนเลส (แท้)": 32.50,
            "ขวดเบียร์ (ช้าง, ลีโอ)": 13.00,
            "เศษแก้วขาว": 1.50
        }

        try:
            # ตรวจสอบว่าไฟล์มีอยู่หรือไม่
            if os.path.exists(prices_file):
                print(f"📁 พบไฟล์ราคา: {prices_file}")

                # ตรวจสอบขนาดไฟล์
                file_size = os.path.getsize(prices_file)
                if file_size == 0:
                    print("⚠️ ไฟล์ราคาว่างเปล่า")
                    raise json.JSONDecodeError("Empty file", "", 0)

                print(f"📊 ขนาดไฟล์: {file_size} bytes")

                # อ่านไฟล์
                with open(prices_file, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    print(f"📄 เนื้อหาไฟล์: {content[:100]}..." if len(
                        content) > 100 else f"📄 เนื้อหาไฟล์: {content}")

                    if not content:
                        raise json.JSONDecodeError("Empty content", "", 0)

                    prices_data = json.loads(content)

                # ตรวจสอบโครงสร้างข้อมูล
                if not isinstance(prices_data, dict):
                    raise ValueError("ไฟล์ราคาต้องเป็น dictionary")

                if 'BUY_PRICES' not in prices_data or 'SELL_PRICES' not in prices_data:
                    print("⚠️ โครงสร้างไฟล์ไม่ถูกต้อง กำลังสร้างใหม่...")
                    raise KeyError("Missing BUY_PRICES or SELL_PRICES")

                # ตรวจสอบว่าข้อมูลราคาเป็น dict และมีข้อมูล
                if not isinstance(prices_data['BUY_PRICES'], dict) or not prices_data['BUY_PRICES']:
                    print("⚠️ BUY_PRICES ไม่ถูกต้องหรือว่างเปล่า")
                    raise ValueError("Invalid BUY_PRICES")

                if not isinstance(prices_data['SELL_PRICES'], dict) or not prices_data['SELL_PRICES']:
                    print("⚠️ SELL_PRICES ไม่ถูกต้องหรือว่างเปล่า")
                    raise ValueError("Invalid SELL_PRICES")

                # โหลดราคา
                self.BUY_PRICES = prices_data['BUY_PRICES']
                self.SELL_PRICES = prices_data['SELL_PRICES']

                print(f"✅ โหลดราคาสำเร็จ:")
                print(f"   - ราคารับซื้อ: {list(self.BUY_PRICES.keys())}")
                print(f"   - ราคาจำหน่าย: {list(self.SELL_PRICES.keys())}")

            else:
                print(f"❌ ไม่พบไฟล์ราคา: {prices_file}")
                raise FileNotFoundError(f"ไม่พบไฟล์ {prices_file}")

        except (FileNotFoundError, json.JSONDecodeError, KeyError, ValueError) as e:
            print(f"⚠️ เกิดปัญหาในการโหลดราคา: {e}")
            print("🔄 กำลังสร้างไฟล์ราคาใหม่...")

            # ใช้ราคาเริ่มต้น
            self.BUY_PRICES = default_buy_prices.copy()
            self.SELL_PRICES = {k: round(v * 1.1, 2)
                                for k, v in self.BUY_PRICES.items()}

            # บันทึกไฟล์ราคาใหม่
            try:
                self.save_prices()
                print(f"✅ สร้างไฟล์ราคาใหม่สำเร็จ: {prices_file}")
                messagebox.showinfo(
                    "สร้างไฟล์ราคาใหม่",
                    f"ไม่พบไฟล์ราคาหรือไฟล์เสียหาย\n"
                    f"สร้างไฟล์ราคาใหม่ที่ {prices_file}\n"
                    f"โหลดราคาเริ่มต้น {len(self.BUY_PRICES)} รายการ"
                )
            except Exception as save_error:
                print(f"❌ ไม่สามารถสร้างไฟล์ราคาใหม่ได้: {save_error}")
                messagebox.showerror(
                    "เกิดข้อผิดพลาด",
                    f"ไม่สามารถสร้างไฟล์ราคาได้: {save_error}\n"
                    f"โปรแกรมจะใช้ราคาเริ่มต้นในหน่วยความจำ"
                )

        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดไม่คาดคิด: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถโหลดราคาได้: {e}")

            # ใช้ราคาเริ่มต้นในกรณีฉุกเฉิน
            self.BUY_PRICES = default_buy_prices.copy()
            self.SELL_PRICES = {k: round(v * 1.1, 2)
                                for k, v in self.BUY_PRICES.items()}

    def save_prices(self):
        """บันทึกราคาลงไฟล์ JSON พร้อมการจัดการข้อผิดพลาด"""
        prices_file = 'prices.json'

        try:
            prices_data = {
                'BUY_PRICES': self.BUY_PRICES,
                'SELL_PRICES': self.SELL_PRICES
            }

            # สร้างไฟล์สำรอง
            if os.path.exists(prices_file):
                backup_file = f"{prices_file}.backup"
                try:
                    import shutil
                    shutil.copy2(prices_file, backup_file)
                    print(f"📋 สร้างไฟล์สำรองที่ {backup_file}")
                except:
                    pass

            # บันทึกไฟล์ใหม่
            with open(prices_file, 'w', encoding='utf-8') as f:
                json.dump(prices_data, f, ensure_ascii=False, indent=4)

            print(f"✅ บันทึกราคาสำเร็จ: {prices_file}")

            # ตรวจสอบไฟล์ที่บันทึก
            with open(prices_file, 'r', encoding='utf-8') as f:
                test_load = json.load(f)
                print(
                    f"🔍 ตรวจสอบไฟล์: BUY_PRICES={len(test_load['BUY_PRICES'])}, SELL_PRICES={len(test_load['SELL_PRICES'])}")

        except Exception as e:
            print(f"❌ ไม่สามารถบันทึกราคาได้: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ราคาได้: {e}")

    def register_thai_font(self):
        """ลงทะเบียนฟอนต์ภาษาไทยสำหรับ PDF"""
        try:
            # ลำดับการค้นหาไฟล์ฟอนต์
            font_paths = [
                os.path.join(os.path.dirname(os.path.abspath(
                    __file__)), "TH Sarabun New Bold.ttf"),
                os.path.join(os.path.dirname(os.path.abspath(
                    __file__)), "THSarabunNew Bold.ttf"),
                os.path.join(os.path.dirname(os.path.abspath(
                    __file__)), "fonts", "TH Sarabun New Bold.ttf"),
                "C:/Windows/Fonts/THSarabunNew Bold.ttf",
                "C:/Windows/Fonts/TH Sarabun New Bold.ttf",
                "/usr/share/fonts/truetype/thai/TH Sarabun New Bold.ttf",
                "/usr/local/share/fonts/TH Sarabun New Bold.ttf",
                "/System/Library/Fonts/TH Sarabun New Bold.ttf",
                "/Library/Fonts/TH Sarabun New Bold.ttf"
            ]

            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('THSarabun', font_path))
                        print(f"✅ ลงทะเบียนฟอนต์สำเร็จ: {font_path}")
                        return 'THSarabun'
                    except Exception as e:
                        print(f"❌ ไม่สามารถลงทะเบียนฟอนต์ {font_path}: {e}")
                        continue

            messagebox.showwarning(
                "ไม่พบฟอนต์ภาษาไทย",
                "ไม่พบไฟล์ฟอนต์ TH Sarabun New Bold.ttf\n\n"
                "กรุณาดาวน์โหลดฟอนต์จาก https://fonts.google.com/specimen/Sarabun\n"
                "แล้ววางไฟล์ในโฟลเดอร์เดียวกับโปรแกรม\n\n"
                "ตอนนี้จะใช้ฟอนต์ Helvetica แทน (อาจแสดงภาษาไทยไม่ถูกต้อง)"
            )
            return 'Helvetica'

        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการลงทะเบียนฟอนต์: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถลงทะเบียนฟอนต์ได้: {e}")
            return 'Helvetica'

    def _setup_transaction_tab(self, tab, mode, prices, update_price_cmd, label1_text, label2_text):
        vcmd = (self.root.register(self.validate_numeric), '%P')

        # Create a frame for the input form
        input_frame = ctk.CTkFrame(tab, fg_color="transparent")
        input_frame.pack(side=ctk.TOP, fill=ctk.X, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text=label1_text, font=("TH Sarabun New", 18)).grid(
            row=0, column=0, padx=10, pady=10, sticky="w")
        name1_var = tk.StringVar()
        ctk.CTkEntry(input_frame, textvariable=name1_var, width=300, font=(
            "TH Sarabun New", 18)).grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text=label2_text, font=("TH Sarabun New", 18)).grid(
            row=1, column=0, padx=10, pady=10, sticky="w")
        name2_var = tk.StringVar()
        ctk.CTkEntry(input_frame, textvariable=name2_var, width=300, font=(
            "TH Sarabun New", 18)).grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="เลือกสินค้า:", font=("TH Sarabun New", 18)).grid(
            row=2, column=0, padx=10, pady=10, sticky="w")
        item_var = tk.StringVar()

        # ตรวจสอบว่ามีราคาให้โหลดหรือไม่
        price_list = list(prices.keys()) if prices else ["ไม่มีข้อมูลราคา"]
        print(f"🏷️ รายการสินค้าสำหรับ {mode}: {price_list}")

        item_combobox = ctk.CTkComboBox(
            input_frame, variable=item_var, values=price_list, width=300, font=("TH Sarabun New", 18), command=update_price_cmd)
        item_combobox.grid(row=2, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="ราคา/กก. (บาท):", font=("TH Sarabun New",
                     18)).grid(row=3, column=0, padx=10, pady=10, sticky="w")
        price_var = tk.DoubleVar()
        ctk.CTkEntry(input_frame, textvariable=price_var, width=300, validate="key", validatecommand=vcmd, font=(
            "TH Sarabun New", 18)).grid(row=3, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="น้ำหนัก (กก.):", font=(
            "TH Sarabun New", 18)).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        weight_var = tk.DoubleVar()
        ctk.CTkEntry(input_frame, textvariable=weight_var, width=300, validate="key", validatecommand=vcmd, font=(
            "TH Sarabun New", 18)).grid(row=4, column=1, padx=10, pady=10, sticky="ew")

        button_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        button_frame.grid(row=5, column=0, columnspan=2, pady=20)

        ctk.CTkButton(button_frame, text="🧮 คำนวณ", command=lambda: self._calculate(
            mode), font=("TH Sarabun New", 18, "bold")).grid(row=0, column=0, padx=20)
        save_print_button = ctk.CTkButton(button_frame, text="💾🖨️ บันทึกและพิมพ์", command=lambda: self._save_print(
            mode), state="disabled", font=("TH Sarabun New", 18, "bold"))
        save_print_button.grid(row=0, column=1, padx=20)

        result_label = ctk.CTkLabel(input_frame, text="กรอกข้อมูลแล้วกดคำนวณ", font=(
            "TH Sarabun New", 22, "bold"), text_color="green")
        result_label.grid(row=6, column=0, columnspan=2, pady=10)

        # --- History Table for this tab ---
        # Create a frame for the history table, so we can pack it to fill the remaining space
        history_frame = ctk.CTkFrame(tab, fg_color="transparent")
        history_frame.pack(side=ctk.BOTTOM, fill=ctk.BOTH,
                           expand=True, padx=10, pady=10)

        if mode == 'in':
            ctk.CTkLabel(history_frame, text="📊 สินค้ารับเข้า", font=(
                "TH Sarabun New", 20, "bold")).pack(anchor="w")
            self.incoming_calc_tree = ttk.Treeview(history_frame, columns=(
                "date", "seller", "buyer", "item", "price", "weight", "total"), show="headings", height=10)
            for col, text in zip(("date", "seller", "buyer", "item", "price", "weight", "total"), ("วันที่", "ผู้ขาย", "ผู้รับ", "สินค้า", "ราคา/กก.", "น้ำหนัก", "รวม")):
                self.incoming_calc_tree.heading(col, text=text)
                self.incoming_calc_tree.column(col, width=150, anchor="center")
            self.incoming_calc_tree.pack(fill=ctk.BOTH, expand=True, pady=5)
            self.seller_var, self.buyer_var = name1_var, name2_var
            self.item_in_var, self.price_in_var, self.weight_in_var = item_var, price_var, weight_var
            self.save_print_in_button, self.result_in_label = save_print_button, result_label
        else:  # mode == 'out'
            ctk.CTkLabel(history_frame, text="📊 สินค้าจำหน่ายออก", font=(
                "TH Sarabun New", 20, "bold")).pack(anchor="w")
            self.outgoing_calc_tree = ttk.Treeview(history_frame, columns=(
                "date", "payer", "recipient", "item", "price", "weight", "total"), show="headings", height=10)
            for col, text in zip(("date", "payer", "recipient", "item", "price", "weight", "total"), ("วันที่", "ผู้จ่าย", "ผู้รับ", "สินค้า", "ราคา/กก.", "น้ำหนัก", "รวม")):
                self.outgoing_calc_tree.heading(col, text=text)
                self.outgoing_calc_tree.column(col, width=150, anchor="center")
            self.outgoing_calc_tree.pack(fill=ctk.BOTH, expand=True, pady=5)
            self.payer_var, self.recipient_var = name1_var, name2_var
            self.item_out_var, self.price_out_var, self.weight_out_var = item_var, price_var, weight_var
            self.save_print_out_button, self.result_out_label = save_print_button, result_label

        # ตั้งค่าค่าเริ่มต้นของสินค้าและราคา
        if prices and len(prices) > 0:
            first_item = list(prices.keys())[0]
            item_var.set(first_item)
            print(f"🎯 ตั้งค่าสินค้าเริ่มต้นเป็น: {first_item}")
            # เรียกใช้ update price command เพื่อตั้งราคา
            # เรียกหลังจาก GUI โหลดเสร็จ
            self.root.after(100, update_price_cmd)
        else:
            print("⚠️ ไม่มีราคาให้ตั้งค่าเริ่มต้น")

    def setup_history_tab(self):
        tables_frame = ctk.CTkFrame(self.history_tab, fg_color="transparent")
        tables_frame.pack(fill=ctk.BOTH, expand=True)
        tables_frame.grid_rowconfigure((1, 3, 5), weight=1)
        tables_frame.grid_columnconfigure(0, weight=1)

        # Receipt In History
        ctk.CTkLabel(tables_frame, text="🧾 ประวัติใบเสร็จรับเข้า", font=(
            "TH Sarabun New", 20, "bold")).grid(row=0, column=0, sticky="w", pady=(10, 0))
        self.receipt_in_tree = ttk.Treeview(tables_frame, columns=(
            "date", "seller", "buyer", "item", "price", "weight", "total", "file"), show="headings", height=6)
        for col, text in zip(("date", "seller", "buyer", "item", "price", "weight", "total", "file"), ("วันที่", "ผู้ขาย", "ผู้รับ", "สินค้า", "ราคา/กก.", "น้ำหนัก", "รวม", "ไฟล์ PDF")):
            self.receipt_in_tree.heading(col, text=text)
            self.receipt_in_tree.column(col, width=150, anchor="center")
        self.receipt_in_tree.grid(row=1, column=0, sticky="nsew", pady=5)

        # Receipt Out History
        ctk.CTkLabel(tables_frame, text="🧾 ประวัติใบเสร็จจำหน่ายออก", font=(
            "TH Sarabun New", 20, "bold")).grid(row=2, column=0, sticky="w", pady=(20, 0))
        self.receipt_out_tree = ttk.Treeview(tables_frame, columns=(
            "date", "payer", "recipient", "item", "price", "weight", "total", "file"), show="headings", height=6)
        for col, text in zip(("date", "payer", "recipient", "item", "price", "weight", "total", "file"), ("วันที่", "ผู้จ่าย", "ผู้รับ", "สินค้า", "ราคา/กก.", "น้ำหนัก", "รวม", "ไฟล์ PDF")):
            self.receipt_out_tree.heading(col, text=text)
            self.receipt_out_tree.column(col, width=150, anchor="center")
        self.receipt_out_tree.grid(row=3, column=0, sticky="nsew", pady=5)

        # Inventory
        ctk.CTkLabel(tables_frame, text="📦 สินค้าคงคลัง", font=(
            "TH Sarabun New", 20, "bold")).grid(row=4, column=0, sticky="w", pady=(20, 0))
        self.inventory_tree = ttk.Treeview(tables_frame, columns=(
            "item", "total_in", "total_out", "stock"), show="headings", height=6)
        for col, text in zip(("item", "total_in", "total_out", "stock"), ("สินค้า", "รวมรับเข้า", "รวมจำหน่ายออก", "คงคลัง")):
            self.inventory_tree.heading(col, text=text)
            self.inventory_tree.column(col, width=150, anchor="center")
        self.inventory_tree.grid(row=5, column=0, sticky="nsew", pady=5)

    def validate_numeric(self, new_value):
        if new_value == "":
            return True
        try:
            float(new_value)
            return True
        except ValueError:
            return False

    def update_buy_price(self, event=None):
        """อัปเดตราคารับซื้อ"""
        try:
            selected_item = self.item_in_var.get()
            print(f"🔄 เลือกสินค้ารับซื้อ: {selected_item}")

            if selected_item and selected_item in self.BUY_PRICES:
                price = self.BUY_PRICES[selected_item]
                self.price_in_var.set(price)
                print(f"💰 ตั้งราคารับซื้อ: {price} บาท/กก.")
            else:
                print(f"⚠️ ไม่พบราคาสำหรับสินค้า: {selected_item}")
                self.price_in_var.set(0.0)
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการอัปเดตราคารับซื้อ: {e}")
            self.price_in_var.set(0.0)

    def update_sell_price(self, event=None):
        """อัปเดตราคาจำหน่าย"""
        try:
            selected_item = self.item_out_var.get()
            print(f"🔄 เลือกสินค้าจำหน่าย: {selected_item}")

            if selected_item and selected_item in self.SELL_PRICES:
                price = self.SELL_PRICES[selected_item]
                self.price_out_var.set(price)
                print(f"💰 ตั้งราคาจำหน่าย: {price} บาท/กก.")
            else:
                print(f"⚠️ ไม่พบราคาสำหรับสินค้า: {selected_item}")
                self.price_out_var.set(0.0)
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการอัปเดตราคาจำหน่าย: {e}")
            self.price_out_var.set(0.0)

    def create_excel_file(self, excel_file):
        """สร้างไฟล์ Excel ใหม่"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Records"
            ws.append(["วันที่", "ชื่อผู้ขาย/ผู้จ่าย", "ชื่อผู้รับ",
                      "สินค้า", "ราคา/กก.", "น้ำหนัก (กก.)", "รวม (บาท)"])
            wb.save(excel_file)
            print(f"✅ สร้างไฟล์ Excel ใหม่: {excel_file}")
        except Exception as e:
            print(f"❌ ไม่สามารถสร้างไฟล์ Excel ได้: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถสร้างไฟล์ Excel ได้: {e}")

    def load_excel_history(self, excel_file, tree):
        """โหลดประวัติจากไฟล์ Excel"""
        if os.path.exists(excel_file):
            try:
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                row_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # ตรวจสอบว่าแถวไม่ว่าง
                        tree.insert("", "end", values=row)
                        row_count += 1
                print(f"📊 โหลดประวัติจาก {excel_file}: {row_count} รายการ")
            except Exception as e:
                print(f"❌ ไม่สามารถโหลดประวัติจาก Excel ได้: {e}")
                messagebox.showwarning(
                    "ข้อผิดพลาดในการโหลด", f"ไม่สามารถโหลดประวัติจาก Excel: {e}")

    def load_receipt_history(self):
        """โหลดประวัติใบเสร็จจากไฟล์ JSON"""
        try:
            if os.path.exists(self.receipt_history_file):
                with open(self.receipt_history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)

                in_count = len(history.get('in', []))
                out_count = len(history.get('out', []))

                for record in history.get('in', []):
                    self.receipt_in_tree.insert(
                        "", "end", values=tuple(record))
                for record in history.get('out', []):
                    self.receipt_out_tree.insert(
                        "", "end", values=tuple(record))

                print(
                    f"📋 โหลดประวัติใบเสร็จ: รับเข้า {in_count} รายการ, จำหน่าย {out_count} รายการ")
            else:
                print(
                    f"📁 ไม่พบไฟล์ประวัติใบเสร็จ: {self.receipt_history_file}")
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️ ไม่สามารถโหลดประวัติใบเสร็จได้: {e}")
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการโหลดประวัติใบเสร็จ: {e}")

    def save_receipt_history(self, mode, data, filename):
        """บันทึกประวัติใบเสร็จ"""
        history = {'in': [], 'out': []}
        try:
            if os.path.exists(self.receipt_history_file):
                with open(self.receipt_history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            print("📝 สร้างไฟล์ประวัติใบเสร็จใหม่")

        record = list(data) + [filename]
        history[mode].append(record)

        try:
            with open(self.receipt_history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=4)
            print(f"✅ บันทึกประวัติใบเสร็จสำเร็จ: {mode}")
        except Exception as e:
            print(f"❌ ไม่สามารถบันทึกประวัติใบเสร็จได้: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถบันทึกประวัติใบเสร็จ: {e}")

    def compute_inventory(self):
        """คำนวณสินค้าคงคลัง"""
        stock = {}

        # นับสินค้ารับเข้า
        if os.path.exists(self.incoming_excel):
            try:
                wb_in = load_workbook(self.incoming_excel, read_only=True)
                ws_in = wb_in.active
                for row in ws_in.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    _, _, _, item, _, weight, _ = row
                    if item and weight:
                        if item not in stock:
                            stock[item] = {'in': 0, 'out': 0}
                        stock[item]['in'] += float(weight)
            except Exception as e:
                print(f"❌ ไม่สามารถอ่านไฟล์รับเข้าได้: {e}")

        # นับสินค้าจำหน่ายออก
        if os.path.exists(self.outgoing_excel):
            try:
                wb_out = load_workbook(self.outgoing_excel, read_only=True)
                ws_out = wb_out.active
                for row in ws_out.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    _, _, _, item, _, weight, _ = row
                    if item and weight:
                        if item not in stock:
                            stock[item] = {'in': 0, 'out': 0}
                        stock[item]['out'] += float(weight)
            except Exception as e:
                print(f"❌ ไม่สามารถอ่านไฟล์จำหน่ายออกได้: {e}")

        # อัปเดต inventory tree
        self.inventory_tree.delete(*self.inventory_tree.get_children())
        for item, data in sorted(stock.items()):
            remaining = data['in'] - data['out']
            self.inventory_tree.insert("", "end", values=(
                item,
                f"{data['in']:.2f}",
                f"{data['out']:.2f}",
                f"{remaining:.2f}"
            ))

        print(f"📦 คำนวณสินค้าคงคลัง: {len(stock)} รายการ")

    def _calculate(self, mode):
        """คำนวณยอดรวม"""
        try:
            # เลือกตัวแปรตามโหมด
            if mode == 'in':
                name1, name2, item, price_var, weight_var = (
                    self.seller_var, self.buyer_var, self.item_in_var,
                    self.price_in_var, self.weight_in_var
                )
                result_label = self.result_in_label
                save_print_button = self.save_print_in_button
                tree = self.incoming_calc_tree
            else:  # mode == 'out'
                name1, name2, item, price_var, weight_var = (
                    self.payer_var, self.recipient_var, self.item_out_var,
                    self.price_out_var, self.weight_out_var
                )
                result_label = self.result_out_label
                save_print_button = self.save_print_out_button
                tree = self.outgoing_calc_tree

            # ดึงค่าจากฟอร์ม
            name1_val = name1.get().strip()
            name2_val = name2.get().strip()
            item_val = item.get()
            price_per_kg = price_var.get()
            weight = weight_var.get()

            print(
                f"🧮 กำลังคำนวณ {mode}: {item_val}, {price_per_kg} บาท/กก., {weight} กก.")

            # ตรวจสอบข้อมูล
            if not name1_val or not name2_val or not item_val or price_per_kg <= 0 or weight <= 0:
                missing = []
                if not name1_val:
                    missing.append("ชื่อผู้ขาย/ผู้จ่าย")
                if not name2_val:
                    missing.append("ชื่อผู้รับ")
                if not item_val:
                    missing.append("สินค้า")
                if price_per_kg <= 0:
                    missing.append("ราคา")
                if weight <= 0:
                    missing.append("น้ำหนัก")

                messagebox.showerror(
                    "ข้อมูลไม่ครบถ้วน", f"กรุณากรอกข้อมูลให้ครบถ้วน:\n- {', '.join(missing)}")
                return

            # คำนวณยอดรวม
            total = price_per_kg * weight
            current_data = (
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                name1_val, name2_val, item_val, price_per_kg, weight, total
            )

            # แสดงผลลัพธ์
            result_label.configure(text=f"รวมทั้งสิ้น: {total:,.2f} บาท")
            save_print_button.configure(state="normal")
            tree.insert("", "end", values=current_data)

            # เก็บข้อมูลปัจจุบัน
            if mode == 'in':
                self.current_in_data = current_data
            else:
                self.current_out_data = current_data

            print(f"✅ คำนวณสำเร็จ: {total:,.2f} บาท")

        except ValueError as e:
            print(f"❌ ข้อผิดพลาดข้อมูลตัวเลข: {e}")
            messagebox.showerror(
                "ข้อผิดพลาด", "กรุณากรอกข้อมูลตัวเลขให้ถูกต้อง")
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการคำนวณ: {e}")
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถคำนวณได้: {e}")

    def _save_print(self, mode):
        """บันทึกและพิมพ์ใบเสร็จ"""
        try:
            # เลือกข้อมูลตามโหมด
            if mode == 'in':
                data = self.current_in_data
                excel_file = self.incoming_excel
                tree = self.receipt_in_tree
            else:  # mode == 'out'
                data = self.current_out_data
                excel_file = self.outgoing_excel
                tree = self.receipt_out_tree

            if not data:
                messagebox.showerror("ไม่มีข้อมูล", "กรุณาคำนวณก่อนบันทึก")
                return

            print(f"💾 กำลังบันทึกข้อมูล {mode}...")

            # บันทึกลง Excel
            self.save_excel(data, excel_file)

            # สร้างใบเสร็จ PDF
            filename = self.print_receipt(data, mode)

            if filename:
                # บันทึกประวัติใบเสร็จ
                self.save_receipt_history(mode, data, filename)
                tree.insert("", "end", values=(*data, filename))
                print(f"✅ บันทึกและสร้างใบเสร็จสำเร็จ: {filename}")

            # อัปเดตสินค้าคงคลัง
            self.compute_inventory()

        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการบันทึก: {e}")
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถบันทึกได้: {e}")

    def save_excel(self, data, excel_file):
        """บันทึกข้อมูลลงไฟล์ Excel"""
        try:
            if os.path.exists(excel_file):
                wb = load_workbook(excel_file)
                ws = wb.active
            else:
                self.create_excel_file(excel_file)
                wb = load_workbook(excel_file)
                ws = wb.active

            ws.append(data)
            wb.save(excel_file)
            print(f"✅ บันทึกข้อมูลลง Excel สำเร็จ: {excel_file}")
            messagebox.showinfo(
                "สำเร็จ", f"บันทึกข้อมูลลง {os.path.basename(excel_file)} เรียบร้อยแล้ว")

        except PermissionError:
            print(f"❌ สิทธิ์การเข้าถึงถูกปฏิเสธ: {excel_file}")
            messagebox.showerror("สิทธิ์การเข้าถึงถูกปฏิเสธ",
                                 "ไม่สามารถบันทึกไฟล์ Excel ได้ กรุณาปิดไฟล์หากเปิดอยู่แล้วลองอีกครั้ง")
        except Exception as e:
            print(f"❌ ไม่สามารถบันทึก Excel ได้: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถบันทึก Excel ได้: {e}")

    def print_receipt(self, data, mode):
        """สร้างใบเสร็จ PDF พร้อมฟอนต์ภาษาไทย"""
        try:
            date, name1, name2, item, price_per_kg, weight, total = data
            dt_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            folder = 'receipts_in' if mode == 'in' else 'receipts_out'
            filename = os.path.join(folder, f"receipt_{dt_str}.pdf")

            print(f"🖨️ กำลังสร้างใบเสร็จ: {filename}")

            # สร้าง PDF
            c = canvas.Canvas(filename, pagesize=letter)
            width, height = letter

            try:
                c.setFont(self.thai_font_name, 20)
            except:
                c.setFont('Helvetica', 20)
                print("⚠️ ไม่สามารถใช้ฟอนต์ไทยได้ ใช้ Helvetica แทน")

            title = "ใบเสร็จรับซื้อของเก่า" if mode == 'in' else "ใบเสร็จจำหน่ายของเก่า"
            title_width = c.stringWidth(title, self.thai_font_name, 20)
            c.drawString((width - title_width) / 2, height - 100, title)

            y_position = height - 150
            line_height = 25

            try:
                c.setFont(self.thai_font_name, 16)
            except:
                c.setFont('Helvetica', 16)

            label1 = "ผู้ขาย:" if mode == 'in' else "ผู้จ่าย:"
            lines = [
                f"{label1} {name1}",
                f"ผู้รับ: {name2}",
                f"สินค้า: {item}",
                f"ราคาต่อหน่วย: {price_per_kg:,.2f} บาท/กก.",
                f"น้ำหนัก: {weight:,.2f} กก.",
            ]

            for line in lines:
                c.drawString(100, y_position, line)
                y_position -= line_height

            c.line(100, y_position - 10, width - 100, y_position - 10)
            y_position -= 30

            try:
                c.setFont(self.thai_font_name, 18)
            except:
                c.setFont('Helvetica-Bold', 18)

            total_text = f"รวมทั้งสิ้น: {total:,.2f} บาท"
            c.drawString(100, y_position, total_text)

            c.line(100, y_position - 20, width - 100, y_position - 20)
            y_position -= 40

            try:
                c.setFont(self.thai_font_name, 14)
            except:
                c.setFont('Helvetica', 14)

            c.drawString(100, y_position, f"วันที่: {date}")

            y_position -= 80
            c.drawString(100, y_position,
                         "ลายเซ็นผู้รับ: _____________________")
            c.drawString(350, y_position,
                         "ลายเซ็นผู้จ่าย: _____________________")

            c.save()

            self.open_file(filename)
            messagebox.showinfo(
                "สำเร็จ", f"สร้างใบเสร็จ PDF เรียบร้อยแล้ว\nไฟล์: {filename}")
            return filename

        except Exception as e:
            print(f"❌ ไม่สามารถสร้างใบเสร็จ PDF ได้: {e}")
            messagebox.showerror(
                "เกิดข้อผิดพลาด", f"ไม่สามารถสร้างใบเสร็จ PDF ได้: {e}")
            return None

    def open_file(self, filename):
        """เปิดไฟล์ด้วยโปรแกรมเริ่มต้นของระบบ"""
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(filename)
            elif system == "Darwin":
                os.system(f"open '{filename}'")
            elif system == "Linux":
                os.system(f"xdg-open '{filename}'")
            else:
                messagebox.showinfo(
                    "เปิดไฟล์", f"ไฟล์ {filename} ถูกสร้างแล้ว แต่ไม่สามารถเปิดอัตโนมัติได้บนระบบนี้ กรุณาเปิดด้วยตัวเอง")
        except Exception as e:
            print(f"❌ ไม่สามารถเปิดไฟล์ได้: {e}")
            messagebox.showinfo(
                "เปิดไฟล์", f"ไฟล์ {filename} ถูกสร้างแล้ว กรุณาเปิดด้วยตัวเอง")

    def open_file(self, filename):
        """เปิดไฟล์ด้วยโปรแกรมเริ่มต้นของระบบ"""
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(filename)
            elif system == "Darwin":
                os.system(f"open '{filename}'")
            elif system == "Linux":
                os.system(f"xdg-open '{filename}'")
            else:
                messagebox.showinfo(
                    "เปิดไฟล์", f"ไฟล์ {filename} ถูกสร้างแล้ว แต่ไม่สามารถเปิดอัตโนมัติได้บนระบบนี้ กรุณาเปิดด้วยตัวเอง")
        except Exception as e:
            print(f"❌ ไม่สามารถเปิดไฟล์ได้: {e}")
            messagebox.showinfo(
                "เปิดไฟล์", f"ไฟล์ {filename} ถูกสร้างแล้ว กรุณาเปิดด้วยตัวเอง")

# ฟังก์ชันสำหรับการดีบัก - ตรวจสอบไฟล์ราคา
def debug_prices_file():
    """ตรวจสอบไฟล์ราคาและแสดงข้อมูลเพื่อการดีบัก"""
    prices_file = 'prices.json'

    print("🔍 === การดีบักไฟล์ราคา ===")
    print(f"📁 ไฟล์: {prices_file}")
    print(f"📂 โฟลเดอร์ปัจจุบัน: {os.getcwd()}")
    print(f"📄 ไฟล์มีอยู่: {os.path.exists(prices_file)}")

    if os.path.exists(prices_file):
        file_size = os.path.getsize(prices_file)
        print(f"📊 ขนาดไฟล์: {file_size} bytes")

        try:
            with open(prices_file, 'r', encoding='utf-8') as f:
                content = f.read()
                print(f"📝 เนื้อหาไฟล์ ({len(content)} ตัวอักษร):")
                print("-" * 50)
                print(content)
                print("-" * 50)

                # ลองแปลง JSON
                f.seek(0)
                data = json.load(f)
                print(f"✅ แปลง JSON สำเร็จ")
                print(f"🔑 คีย์ที่พบ: {list(data.keys())}")

                if 'BUY_PRICES' in data:
                    print(
                        f"💰 ราคารับซื้อ: {len(data['BUY_PRICES'])} รายการ")
                    for item, price in list(data['BUY_PRICES'].items())[:3]:
                        print(f"   - {item}: {price}")
                    if len(data['BUY_PRICES']) > 3:
                        print(
                            f"   ... และอีก {len(data['BUY_PRICES']) - 3} รายการ")

                if 'SELL_PRICES' in data:
                    print(
                        f"🏪 ราคาจำหน่าย: {len(data['SELL_PRICES'])} รายการ")
                    for item, price in list(data['SELL_PRICES'].items())[:3]:
                        print(f"   - {item}: {price}")
                    if len(data['SELL_PRICES']) > 3:
                        print(
                            f"   ... และอีก {len(data['SELL_PRICES']) - 3} รายการ")

        except json.JSONDecodeError as e:
            print(f"❌ ไฟล์ JSON ผิดรูปแบบ: {e}")
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการอ่านไฟล์: {e}")
    else:
        print("❌ ไม่พบไฟล์ราคา")

    print("🔍 === จบการดีบัก ===\n")

if __name__ == "__main__":
    # เรียกใช้ฟังก์ชันดีบักก่อนเริ่มโปรแกรม
    debug_prices_file()

    print("🚀 เริ่มต้นโปรแกรม...")
    root = ctk.CTk()
    app = ScrapShopApp(root)
    root.mainloop()
